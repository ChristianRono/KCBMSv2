import pandas as pd
import openpyxl as xl

from master.models import Application,FinancialYear

def file_handler2(filename,allocation):
    data = pd.read_excel(filename)
    errors = []
    if not data.isnull().values.any():
        if not data['BIRTH CERTIFICATE NUMBER'].duplicated().any():
            if data.loc[:,'AMOUNT'].sum() < allocation:
                return errors,data
            else:
                errors.append('Amount disbursed is more than the amount allocated!')
        else:
            errors.append('There are duplicate Birth Certificate Numbers!')
    else:
        errors.append('There are empty cells in the document! Make sure all data is entered for each student')
    return errors,None

def file_creator(applications):
    wb = xl.Workbook()
    ws = wb.active
    c1 = ws.cell(row=1,column=1)
    c1.value = 'Full Name'
    c2 = ws.cell(row=1,column=2)
    c2.value = 'Admission/Registration Number'
    c3 = ws.cell(row=1,column=3)
    c3.value = 'Gender'
    c4 = ws.cell(row=1,column=4)
    c4.value = 'Amount'

    i = 2
    for data in applications:
        c1 = ws.cell(row=i,column=1)
        c1.value = data.full_name
        c2 = ws.cell(row=i,column=2)
        c2.value = data.admission_no
        c3 = ws.cell(row=i,column=3)
        c3.value = data.get_gender_display()
        c4 = ws.cell(row=i,column=4)
        c4.value = data.amount
        i += 1
    
    wb.save('Students List.xlsx')